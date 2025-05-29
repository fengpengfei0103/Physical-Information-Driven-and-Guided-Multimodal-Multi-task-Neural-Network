# 本代码用于固定入渗情况下，时间变量引导下的边坡饱和度、渗透压、以及稳定性全过程全面预测
# 输入：input1：初始边坡非饱和状态图，
#      input2：入渗时间T（或者说是降雨时间T），
#      input3：初始边坡(应力/孔隙压力)图，
#      input4：初始边坡(应力/孔隙压力)图，
#      input5：初始边坡位移变形图
# 输出：output1：入渗时间T（或者说是降雨时间T）后的边坡饱和状态图像，
#      output2：入渗时间T（或者说是降雨时间T）后的边坡(应力/孔隙压力)图像
#      output3：入渗时间T（或者说是降雨时间T）后的边坡(应力/孔隙压力)图像
#      output4：入渗时间T（或者说是降雨时间T）后的边坡位移变形图像
#      output5：入渗时间T（或者说是降雨时间T）后的边坡稳定性
# Time：2024-12-26
# Email：fpf0103@163.com & 571428374@qq.com

import torch
import torch.nn as nn
import torch.nn.functional as F
from torch.utils.data import DataLoader, Dataset
from torchvision import transforms
from PIL import Image
import numpy as np
import os
import matplotlib.pyplot as plt
import pandas as pd
from torchmetrics.image import PeakSignalNoiseRatio, StructuralSimilarityIndexMeasure
from torchmetrics.regression import MeanSquaredError, MeanAbsoluteError, R2Score
from einops import rearrange
import numbers
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import copy

# 显卡以及CUDA是否可用
if torch.cuda.is_available():
    print("GPU version installed.")
    print(f"CUDA version: {torch.version.cuda}")
    print(f"Available GPUs: {torch.cuda.device_count()}")
    print(f"Current GPU: {torch.cuda.get_device_name(torch.cuda.current_device())}")
else:
    print("CPU version installed.")

physics_params = {
    'S_r': 0, 'S_s': 1.0,    # 饱和度参数
    'alpha_vg': 1.0, 'n': 2.0, 'l': 0.5,  # VG模型参数
    'K_s': 7e-5, 'mu': 1e-3,    # 渗透参数
    'rho': 1000, 'g': 9.81,      # 流体参数
    'E': 1e8, 'nu': 0.3,        # 弹性参数
    'alpha': 1.0,               # Biot系数
    'phi': 0.3                  # 孔隙率
}
# 数据集定义
class SlopeDataset(Dataset):
    def __init__(self, x1_path, x2_values, x3_path, x4_path, y1_path, y2_path, y3_path, y5_values, transform=None):
        self.x1_files = [os.path.join(x1_path, f) for f in os.listdir(x1_path) if f.endswith('.png')]
        self.x2_values = x2_values
        self.x3_files = [os.path.join(x3_path, f) for f in os.listdir(x3_path) if f.endswith('.png')]
        self.x4_files = [os.path.join(x4_path, f) for f in os.listdir(x4_path) if f.endswith('.png')]
        self.y1_files = [os.path.join(y1_path, f) for f in os.listdir(y1_path) if f.endswith('.png')]
        self.y2_files = [os.path.join(y2_path, f) for f in os.listdir(y2_path) if f.endswith('.png')]
        self.y3_files = [os.path.join(y3_path, f) for f in os.listdir(y3_path) if f.endswith('.png')]
        self.y5_values = y5_values
        self.transform = transform

    def __len__(self):
        return len(self.x1_files)

    def __getitem__(self, idx):
        x1 = Image.open(self.x1_files[idx]).convert('RGB')
        x2 = self.x2_values[idx]
        x3 = Image.open(self.x3_files[idx]).convert('RGB')
        x4 = Image.open(self.x4_files[idx]).convert('RGB')
        y1 = Image.open(self.y1_files[idx]).convert('RGB')
        y2 = Image.open(self.y2_files[idx]).convert('RGB')
        y3 = Image.open(self.y3_files[idx]).convert('RGB')
        y5 = self.y5_values[idx]

        if self.transform:
            x1 = self.transform(x1)
            x3 = self.transform(x3)
            x4 = self.transform(x4)
            y1 = self.transform(y1)
            y2 = self.transform(y2)
            y3 = self.transform(y3)

        return x1, x2, x3, x4, y1, y2, y3, y5

# 数据预处理
transform = transforms.Compose([
    transforms.ToTensor(),
])


# 基本模块定义
class LayerNorm(nn.Module):
    r""" LayerNorm that supports two data formats: channels_last (default) or channels_first.
    The ordering of the dimensions in the inputs. channels_last corresponds to inputs with
    shape (batch_size, height, width, channels) while channels_first corresponds to inputs
    with shape (batch_size, channels, height, width).
    """

    def __init__(self, normalized_shape, eps=1e-6, data_format="channels_last"):
        super().__init__()
        self.weight = nn.Parameter(torch.ones(normalized_shape))
        self.bias = nn.Parameter(torch.zeros(normalized_shape))
        self.eps = eps
        self.data_format = data_format
        if self.data_format not in ["channels_last", "channels_first"]:
            raise NotImplementedError
        self.normalized_shape = (normalized_shape,)

    def forward(self, x):
        if self.data_format == "channels_last":
            return F.layer_norm(x, self.normalized_shape, self.weight, self.bias, self.eps)
        elif self.data_format == "channels_first":
            u = x.mean(1, keepdim=True)
            s = (x - u).pow(2).mean(1, keepdim=True)
            x = (x - u) / torch.sqrt(s + self.eps)
            x = self.weight[:, None, None] * x + self.bias[:, None, None]
            return x
class GSAU(nn.Module):
    def __init__(self, n_feats, drop=0.0, k=2, squeeze_factor=15, attn='GLKA'):
        super().__init__()
        i_feats = n_feats * 2

        self.Conv1 = nn.Conv2d(n_feats, i_feats, 1, 1, 0)
        self.DWConv1 = nn.Conv2d(n_feats, n_feats, 7, 1, 7 // 2, groups=n_feats)
        self.Conv2 = nn.Conv2d(n_feats, n_feats, 1, 1, 0)

        self.norm = LayerNorm(n_feats, data_format='channels_first')
        self.scale = nn.Parameter(torch.zeros((1, n_feats, 1, 1)), requires_grad=True)

    def forward(self, x):
        shortcut = x.clone()

        # Ghost Expand
        x = self.Conv1(self.norm(x))
        a, x = torch.chunk(x, 2, dim=1)
        x = x * self.DWConv1(a)
        x = self.Conv2(x)

        return x * self.scale + shortcut

class MLKA(nn.Module):
    def __init__(self, n_feats):
        super().__init__()
        i_feats = 2 * n_feats
        self.n_feats = n_feats
        self.i_feats = i_feats

        self.norm = LayerNorm(n_feats, data_format='channels_first')
        self.scale = nn.Parameter(torch.zeros((1, n_feats, 1, 1)), requires_grad=True)

        # Multiscale Large Kernel Attention
        self.LKA7 = nn.Sequential(
            nn.Conv2d(n_feats // 4, n_feats // 4, 7, 1, 7 // 2, groups=n_feats // 4),
            nn.Conv2d(n_feats // 4, n_feats // 4, 9, stride=1, padding=(9 // 2) * 4, groups=n_feats // 4, dilation=4),
            nn.Conv2d(n_feats // 4, n_feats // 4, 1, 1, 0))
        self.LKA5 = nn.Sequential(
            nn.Conv2d(n_feats // 4, n_feats // 4, 5, 1, 5 // 2, groups=n_feats // 4),
            nn.Conv2d(n_feats // 4, n_feats // 4, 7, stride=1, padding=(7 // 2) * 3, groups=n_feats // 4, dilation=3),
            nn.Conv2d(n_feats // 4, n_feats // 4, 1, 1, 0))
        self.LKA3 = nn.Sequential(
            nn.Conv2d(n_feats // 4, n_feats // 4, 3, 1, 1, groups=n_feats // 4),
            nn.Conv2d(n_feats // 4, n_feats // 4, 5, stride=1, padding=(5 // 2) * 2, groups=n_feats // 4, dilation=2),
            nn.Conv2d(n_feats // 4, n_feats // 4, 1, 1, 0))
        self.LKA1 = nn.Sequential(
            nn.Conv2d(n_feats // 4, n_feats // 4, 1, 1, 0, groups=n_feats // 4),
            nn.Conv2d(n_feats // 4, n_feats // 4, 3, stride=1, padding=(3 // 2) * 2, groups=n_feats // 4, dilation=2),
            nn.Conv2d(n_feats // 4, n_feats // 4, 1, 1, 0))

        self.X1 = nn.Conv2d(n_feats // 4, n_feats // 4, 1, 1, 0, groups=n_feats // 4)
        self.X3 = nn.Conv2d(n_feats // 4, n_feats // 4, 3, 1, 1, groups=n_feats // 4)
        self.X5 = nn.Conv2d(n_feats // 4, n_feats // 4, 5, 1, 5 // 2, groups=n_feats // 4)
        self.X7 = nn.Conv2d(n_feats // 4, n_feats // 4, 7, 1, 7 // 2, groups=n_feats // 4)

        self.proj_first = nn.Sequential(
            nn.Conv2d(n_feats, i_feats, 1, 1, 0))

        self.proj_last = nn.Sequential(
            nn.Conv2d(n_feats, n_feats, 1, 1, 0))

    def forward(self, x):
        shortcut = x.clone()
        x = self.norm(x)
        x = self.proj_first(x)
        a, x = torch.chunk(x, 2, dim=1)
        a_1, a_2, a_3, a_4 = torch.chunk(a, 4, dim=1)
        a = torch.cat([self.LKA1(a_1) * self.X1(a_1), self.LKA3(a_2) * self.X3(a_2), self.LKA5(a_3) * self.X5(a_3),
                       self.LKA7(a_4) * self.X7(a_4)], dim=1)
        x = self.proj_last(x * a) * self.scale + shortcut
        return x
 # MAB
class MAB(nn.Module):
    def __init__(self, n_feats):
        super().__init__()
        self.LKA = MLKA(n_feats)
        self.LFE = GSAU(n_feats)
    def forward(self, x):
        # large kernel attention
        x = self.LKA(x)
        # local feature extraction
        x = self.LFE(x)
        return x

norm_dict = {'BATCH': nn.BatchNorm2d, 'INSTANCE': nn.InstanceNorm2d, 'GROUP': nn.GroupNorm, 'LAYER': nn.LayerNorm}
class Identity(nn.Module):
    """
    Identity mapping for building a residual connection
    """
    def __init__(self):
        super().__init__()

    def forward(self, x):
        return x
class ConvNorm(nn.Module):
    """
    Convolution and normalization
    """
    def __init__(self, in_channels, out_channels, kernel_size, stride=1, leaky=True, norm='BATCH', activation=True):
        super().__init__()
        # determine basic attributes
        self.norm_type = norm
        padding = (kernel_size - 1) // 2

        # activation, support PReLU and common ReLU
        if activation:
            self.act = nn.PReLU() if leaky else nn.ReLU(inplace=True)
            # self.act = nn.ELU() if leaky else nn.ReLU(inplace=True)
        else:
            self.act = None

        # instantiate layers
        self.conv = nn.Conv2d(in_channels, out_channels, kernel_size, stride, padding, bias=False)
        norm_layer = norm_dict[norm]
        if norm in ['BATCH', 'INSTANCE', 'LAYER']:
            self.norm = norm_layer(out_channels)
        else:
            self.norm = norm_layer(4, in_channels)

    def basic_forward(self, x):
        x = self.conv(x)
        x = self.norm(x)
        if self.act:
            x = self.act(x)
        return x

    def group_forward(self, x):
        x = self.norm(x)
        if self.act:
            x = self.act(x)
        x = self.conv(x)
        return x

    def forward(self, x):
        if self.norm_type in ['BATCH', 'INSTANCE', 'LAYER']:
            return self.basic_forward(x)
        else:
            return self.group_forward(x)

class ResBlock(nn.Module):
    """
    Residual blocks
    """
    def __init__(self, in_channels, out_channels, stride=1, use_dropout=False, leaky=False, norm='INSTANCE'):
        super().__init__()
        self.norm_type = norm
        self.act = nn.PReLU() if leaky else nn.ReLU(inplace=True)
        self.dropout = nn.Dropout3d(p=0.1) if use_dropout else None
        # self.act = nn.ELU() if leaky else nn.ReLU(inplace=True)

        self.conv1 = ConvNorm(in_channels, out_channels, 3, stride, leaky, norm, True)
        self.conv2 = ConvNorm(out_channels, out_channels, 3, 1, leaky, norm, False)

        need_map = in_channels != out_channels or stride != 1
        self.id = ConvNorm(in_channels, out_channels, 1, stride, leaky, norm, False) if need_map else Identity()

    def forward(self, x):
        identity = x
        out = self.conv1(x)
        out = self.conv2(out)
        identity = self.id(identity)

        out = out + identity
        if self.norm_type != 'GROUP':
            out = self.act(out)

        if self.dropout:
            out = self.dropout(out)
        return out

##########################################################################
## Layer Norm

def to_3d(x):
    return rearrange(x, 'b c h w -> b (h w) c')

def to_4d(x,h,w):
    return rearrange(x, 'b (h w) c -> b c h w',h=h,w=w)
class BiasFree_LayerNorm(nn.Module):
    def __init__(self, normalized_shape):
        super(BiasFree_LayerNorm, self).__init__()
        if isinstance(normalized_shape, numbers.Integral):
            normalized_shape = (normalized_shape,)
        normalized_shape = torch.Size(normalized_shape)

        assert len(normalized_shape) == 1

        self.weight = nn.Parameter(torch.ones(normalized_shape))
        self.normalized_shape = normalized_shape

    def forward(self, x):
        sigma = x.var(-1, keepdim=True, unbiased=False)
        return x / torch.sqrt(sigma + 1e-5) * self.weight


class WithBias_LayerNorm(nn.Module):
    def __init__(self, normalized_shape):
        super(WithBias_LayerNorm, self).__init__()
        if isinstance(normalized_shape, numbers.Integral):
            normalized_shape = (normalized_shape,)
        normalized_shape = torch.Size(normalized_shape)

        assert len(normalized_shape) == 1

        self.weight = nn.Parameter(torch.ones(normalized_shape))
        self.bias = nn.Parameter(torch.zeros(normalized_shape))
        self.normalized_shape = normalized_shape

    def forward(self, x):
        mu = x.mean(-1, keepdim=True)
        sigma = x.var(-1, keepdim=True, unbiased=False)
        # print(f"x shape: {x.shape}, mu shape: {mu.shape}, sigma shape: {sigma.shape}")
        # print(f"weight shape: {self.weight.shape}, bias shape: {self.bias.shape}")

        return (x - mu) / torch.sqrt(sigma + 1e-5) * self.weight + self.bias


class LayerNormTransformer(nn.Module):
    def __init__(self, dim, LayerNorm_type):
        super(LayerNormTransformer, self).__init__()
        if LayerNorm_type == 'BiasFree':
            self.body = BiasFree_LayerNorm(dim)
        else:
            self.body = WithBias_LayerNorm(dim)

    def forward(self, x):
        h, w = x.shape[-2:]
        return to_4d(self.body(to_3d(x)), h, w)


##########################################################################
## Gated-Dconv Feed-Forward Network (GDFN)
class FeedForward(nn.Module):
    def __init__(self, dim, ffn_expansion_factor, bias):
        super(FeedForward, self).__init__()

        hidden_features = int(dim * ffn_expansion_factor)

        self.project_in = nn.Conv2d(dim, hidden_features * 2, kernel_size=1, bias=bias)

        self.dwconv = nn.Conv2d(hidden_features * 2, hidden_features * 2, kernel_size=3, stride=1, padding=1,
                                groups=hidden_features * 2, bias=bias)

        self.project_out = nn.Conv2d(hidden_features, dim, kernel_size=1, bias=bias)

    def forward(self, x):
        x = self.project_in(x)
        x1, x2 = self.dwconv(x).chunk(2, dim=1)
        x = F.gelu(x1) * x2
        x = self.project_out(x)
        return x


##########################################################################
## Multi-DConv Head Transposed Self-Attention (MDTA)
class Attention(nn.Module):
    def __init__(self, dim, num_heads, bias):
        super(Attention, self).__init__()
        self.num_heads = num_heads
        self.temperature = nn.Parameter(torch.ones(num_heads, 1, 1))

        self.qkv = nn.Conv2d(dim, dim * 3, kernel_size=1, bias=bias)
        self.qkv_dwconv = nn.Conv2d(dim * 3, dim * 3, kernel_size=3, stride=1, padding=1, groups=dim * 3, bias=bias)
        self.project_out = nn.Conv2d(dim, dim, kernel_size=1, bias=bias)

    def forward(self, x):
        b, c, h, w = x.shape

        qkv = self.qkv_dwconv(self.qkv(x))
        q, k, v = qkv.chunk(3, dim=1)

        q = rearrange(q, 'b (head c) h w -> b head c (h w)', head=self.num_heads)
        k = rearrange(k, 'b (head c) h w -> b head c (h w)', head=self.num_heads)
        v = rearrange(v, 'b (head c) h w -> b head c (h w)', head=self.num_heads)

        q = torch.nn.functional.normalize(q, dim=-1)
        k = torch.nn.functional.normalize(k, dim=-1)

        attn = (q @ k.transpose(-2, -1)) * self.temperature
        attn = attn.softmax(dim=-1)

        out = (attn @ v)

        out = rearrange(out, 'b head c (h w) -> b (head c) h w', head=self.num_heads, h=h, w=w)

        out = self.project_out(out)
        return out
##########################################################################
## Transformer Block
class TransformerBlock(nn.Module):
    def __init__(self, dim, num_heads, ffn_expansion_factor, bias, LayerNorm_type):
        super(TransformerBlock, self).__init__()

        self.norm1 = LayerNormTransformer(dim, LayerNorm_type)
        self.attn = Attention(dim, num_heads, bias)
        self.norm2 = LayerNormTransformer(dim, LayerNorm_type)
        self.ffn = FeedForward(dim, ffn_expansion_factor, bias)

    def forward(self, x):
        x = x + self.attn(self.norm1(x))
        x = x + self.ffn(self.norm2(x))

        return x

# 时间序列特征模块定义
class FeatureModule(nn.Module):
    def __init__(self):
        super(FeatureModule, self).__init__()
        self.fc = nn.Linear(2, 20 * 35 * 8)
        self.relu = nn.ReLU()
        self.reshape = lambda x: x.view(-1, 8, 20, 35)

        self.ConvNorm1 = ConvNorm(in_channels=8, out_channels=8, kernel_size=3, stride=1, leaky=False, norm='INSTANCE', activation=True)
        self.encoder_level = nn.Sequential(*[
            TransformerBlock(dim=8, num_heads=1, ffn_expansion_factor=2.66, bias=False,
                             LayerNorm_type='WithBias') for i in range(4)])

        self.conv_blocks = nn.Sequential(
            nn.Conv2d(8, 8, kernel_size=3, stride=1, padding=1, bias=False),
            self.encoder_level,
            nn.InstanceNorm2d(8),
            nn.ReLU(),
        )
        self.ConvNorm2 = ConvNorm(in_channels=8, out_channels=16, kernel_size=3, stride=1, leaky=False, norm='INSTANCE',
                                  activation=True)

    def forward(self, x):
        x = self.fc(x)
        x = self.relu(x)
        x = self.reshape(x)
        x = self.ConvNorm1(x)
        # print(x.shape)
        x = self.conv_blocks(x)
        x = self.ConvNorm2(x)
        return x
class FeatureModuleStab(nn.Module):
    def __init__(self):
        super(FeatureModuleStab, self).__init__()
        self.fc = nn.Linear(2, 40 * 70 * 2)
        self.relu = nn.ReLU()
        self.reshape = lambda x: x.view(-1, 2, 40, 70)

        self.ConvNorm1 = ConvNorm(in_channels=2, out_channels=8, kernel_size=3, stride=1,
                                  leaky=False, norm='INSTANCE', activation=True)
        self.encoder_level = nn.Sequential(*[
            TransformerBlock(dim=8, num_heads=1, ffn_expansion_factor=2.66, bias=False,
                             LayerNorm_type='WithBias') for i in range(4)])

        self.conv_blocks = nn.Sequential(
            nn.Conv2d(8, 8, kernel_size=3, stride=1, padding=1, bias=False),
            self.encoder_level,
            nn.InstanceNorm2d(8),
            nn.ReLU(),
        )
        self.ConvNorm2 = ConvNorm(in_channels=8, out_channels=16, kernel_size=3, stride=1,
                                  leaky=False, norm='INSTANCE', activation=True)
        self.fc2 = nn.Sequential(
            nn.Linear(2, 40 * 70 * 2),
            nn.ReLU(),
            nn.Linear(40 * 70 * 2, 2),
        )

    def forward(self, value):
        x = self.fc(value)
        x = self.relu(x)
        x = self.reshape(x)
        x = self.ConvNorm1(x)
        x = self.conv_blocks(x)
        x1 = self.ConvNorm2(x)
        x2 = self.fc2(value)

        return x1,x2

# 孔隙压力初始条件化控制模块定义
class PorePressureControlModule(nn.Module):
    def __init__(self):
        super(PorePressureControlModule, self).__init__()
        # Encoder layers
        self.EncoderConvNorm1 = ConvNorm(in_channels=3, out_channels=16, kernel_size=3, stride=1, leaky=False,
                                         norm='INSTANCE',
                                         activation=True)
        self.EncoderMaxPool = nn.MaxPool2d(2, stride=2)
        # self.EncoderConvNorm2 = ConvNorm(in_channels=16, out_channels=16, kernel_size=3, stride=1, leaky=False, norm='INSTANCE',
        #                           activation=True)
        self.EncoderConvNorm2 = ResBlock(in_channels=16, out_channels=16, stride=1, leaky=False,
                                         norm='INSTANCE')
        self.MAB = MAB(16)

        # Bridge layers
        self.bridge = nn.Sequential(
            nn.Conv2d(16, 16, kernel_size=3, stride=1, padding=1),
            nn.ReLU(),
            nn.Conv2d(16, 16, kernel_size=3, stride=1, padding=1),
            nn.ReLU(),
            nn.Dropout(0.5)
        )

    def forward(self, x):
        x = self.EncoderConvNorm1(x)
        x1 = self.EncoderMaxPool(x)
        x = self.EncoderConvNorm2(x1)
        x = self.MAB(x)
        x2 = self.EncoderMaxPool(x)
        x = self.EncoderConvNorm2(x2)
        x = self.MAB(x)
        x3 = self.EncoderMaxPool(x)
        x = self.EncoderConvNorm2(x3)
        x = self.MAB(x)
        x4 = self.EncoderMaxPool(x)

        x = self.bridge(x4)
        return x

# 应力初始条件化控制模块定义
class StressControlModule(nn.Module):
    def __init__(self):
        super(StressControlModule, self).__init__()
        # Encoder layers
        self.EncoderConvNorm1 = ConvNorm(in_channels=3, out_channels=16, kernel_size=3, stride=1, leaky=False,
                                         norm='INSTANCE',
                                         activation=True)
        self.EncoderMaxPool = nn.MaxPool2d(2, stride=2)
        # self.EncoderConvNorm2 = ConvNorm(in_channels=16, out_channels=16, kernel_size=3, stride=1, leaky=False, norm='INSTANCE',
        #                           activation=True)
        self.EncoderConvNorm2 = ResBlock(in_channels=16, out_channels=16, stride=1, leaky=False,
                                         norm='INSTANCE')

        self.MAB = MAB(16)

        # Bridge layers
        self.bridge = nn.Sequential(
            nn.Conv2d(16, 16, kernel_size=3, stride=1, padding=1),
            nn.ReLU(),
            nn.Conv2d(16, 16, kernel_size=3, stride=1, padding=1),
            nn.ReLU(),
            nn.Dropout(0.5)
        )

    def forward(self, x):
        x = self.EncoderConvNorm1(x)
        x1 = self.EncoderMaxPool(x)
        x = self.EncoderConvNorm2(x1)
        x = self.MAB(x)
        x2 = self.EncoderMaxPool(x)
        x = self.EncoderConvNorm2(x2)
        x = self.MAB(x)
        x3 = self.EncoderMaxPool(x)
        x = self.EncoderConvNorm2(x3)
        x = self.MAB(x)
        x4 = self.EncoderMaxPool(x)

        x = self.bridge(x4)
        return x

class CBAM(nn.Module):
    def __init__(self, gate_channels, reduction_ratio=16):
        super(CBAM, self).__init__()

        # 通道注意力
        self.channel_attention = nn.Sequential(
            nn.AdaptiveAvgPool2d(1),  # 全局平均池化
            nn.Conv2d(gate_channels, gate_channels // reduction_ratio, kernel_size=1),
            nn.ReLU(),
            nn.Conv2d(gate_channels // reduction_ratio, gate_channels, kernel_size=1),
            nn.Sigmoid()  # 输出0-1的权重
        )

        # 空间注意力
        self.spatial_attention = nn.Sequential(
            nn.Conv2d(2, 1, kernel_size=7, padding=3),  # 用7x7卷积捕获空间关系
            nn.Sigmoid()
        )

    def forward(self, x):
        # 通道注意力
        channel_att = self.channel_attention(x)
        x_channel_att = x * channel_att  # 通道加权

        # 空间注意力
        avg_out = torch.mean(x_channel_att, dim=1, keepdim=True)  # 通道平均
        max_out, _ = torch.max(x_channel_att, dim=1, keepdim=True)  # 通道最大
        spatial_att_input = torch.cat([avg_out, max_out], dim=1)
        spatial_att = self.spatial_attention(spatial_att_input)
        x_att = x_channel_att * spatial_att  # 空间加权

        return x_att

# 主网络结构
class SlopeNet(nn.Module):
    def __init__(self):
        super(SlopeNet, self).__init__()

        # Encoder layers
        self.EncoderConvNorm1 = ConvNorm(in_channels=3, out_channels=16, kernel_size=3, stride=1, leaky=False, norm='INSTANCE',
                                  activation=True)
        self.EncoderMaxPool = nn.MaxPool2d(2, stride=2)
        # self.EncoderConvNorm2 = ConvNorm(in_channels=16, out_channels=16, kernel_size=3, stride=1, leaky=False, norm='INSTANCE',
        #                           activation=True)
        self.EncoderConvNorm2 = ResBlock(in_channels=16, out_channels=16, stride=1, leaky=False,
                                         norm='INSTANCE')
        self.MAB = MAB(16)

        # Bridge layers
        self.bridge = nn.Sequential(
            nn.Conv2d(16, 16, kernel_size=3, stride=1, padding=1),
            nn.ReLU(),
            nn.Conv2d(16, 16, kernel_size=3, stride=1, padding=1),
            nn.ReLU(),
            nn.Dropout(0.5)
        )

        # Decoder layers
        self.decoderConvTrans = nn.Sequential(
            nn.ConvTranspose2d(16, 16, kernel_size=3, stride=2, padding=1, output_padding=1),
            # nn.BatchNorm2d(16),
            nn.InstanceNorm2d(16),
            nn.ReLU(),
        )
        self.decodeConvNorm1 = ConvNorm(in_channels=16, out_channels=16, kernel_size=3, stride=1, leaky=False,
                                         norm='INSTANCE',
                                         activation=True)

        # Final layer
        self.final_conv = nn.Conv2d(16, 3, kernel_size=3, stride=1, padding=1)
        self.sigmoid = nn.Sigmoid()

        self.FeatureModule = FeatureModule()
        self.FeatureModuleStab = FeatureModuleStab()
        self.PorePressureControlModule  = PorePressureControlModule()
        self.StressControlModule = StressControlModule()

        # 稳定性预测部分架构
        self.stabilyBasicModule = nn.Sequential(
            ConvNorm(in_channels=16, out_channels=32, kernel_size=3, stride=1, leaky=False,
                     norm='INSTANCE',
                     activation=True),
            nn.MaxPool2d(2, stride=2),
            CBAM(32),
            ConvNorm(in_channels=32, out_channels=16, kernel_size=3, stride=1, leaky=False,
                     norm='INSTANCE',
                     activation=True),
            nn.MaxPool2d(2, stride=2),
            CBAM(16),
        )

        # 稳定性部分全连接以及池化层
        self.flatten = nn.Flatten()
        self.fc1 = nn.Linear(16 * 40 * 70, 128)
        self.fc2 = nn.Linear(128, 2)
        self.fc3 = nn.Linear(2, 1)

        # 增量系数
        self.log_phvar1 = nn.Parameter(torch.zeros(1))  # 任务1的噪声参数
        self.log_phvar2 = nn.Parameter(torch.zeros(1))  # 任务2的噪声参数
        self.log_phvar3 = nn.Parameter(torch.zeros(1))  # 任务3的噪声参数
        # 图片任务和稳定性任务
        self.log_var1 = nn.Parameter(torch.zeros(1))  # 任务1的噪声参数
        self.log_var2 = nn.Parameter(torch.zeros(1))  # 任务2的噪声参数

    def compute_Saturability_PorePressure(self,Sr):
        m = 1
        n = 2
        alpha = 1
        waterp = 1000
        g = 9.81
        # 避免除零错误
        S_clamp = torch.clamp(Sr, 0.05, 0.99)

        # Sr = torch.sigmoid(Sr)
        term = torch.pow(S_clamp, -1 / m) - 1  # 计算 Sr^(-1/m) - 1
        term = torch.pow(term, 1 / n)  # 计算 (term)^(1/n)
        psi_prime = (-1 / alpha) * term  # 乘以 -1/alpha
        # psi_prime = torch.sigmoid(waterp * g * psi_prime)

        return waterp * g * psi_prime

    def forward(self, x, featureinit, PorePressure, Stress):
        x = self.EncoderConvNorm1(x)
        x1 = self.EncoderMaxPool(x)
        x = self.EncoderConvNorm2(x1)
        x = self.MAB(x)
        x2 = self.EncoderMaxPool(x)
        x = self.EncoderConvNorm2(x2)
        x = self.MAB(x)
        x3 = self.EncoderMaxPool(x)
        x = self.EncoderConvNorm2(x3)
        x = self.MAB(x)
        x4 = self.EncoderMaxPool(x)

        x = self.bridge(x4)

        # 查看 x 和 feature 的形状
        # print(f"x shape after bridge: {x.shape}")
        feature = self.FeatureModule(featureinit.float())
        PorePressureControl = self.PorePressureControlModule(PorePressure)
        StressControl = self.StressControlModule(Stress)
        # 融合特征
        temp = x * feature
        x = x + temp
        PorePressureControl = (PorePressureControl + PorePressureControl * feature + self.compute_Saturability_PorePressure(
                                      x)) / 2
        StressControl = StressControl + (StressControl * feature + x * PorePressureControl) / 2

        x = self.decoderConvTrans(x)
        x = self.decodeConvNorm1(x)
        feature = self.decoderConvTrans(feature)
        PorePressureControl = self.decoderConvTrans(PorePressureControl)
        PorePressureControl = self.decodeConvNorm1(PorePressureControl)
        StressControl = self.decoderConvTrans(StressControl)
        StressControl = self.decodeConvNorm1(StressControl)


        x = self.decoderConvTrans(x)
        x = self.decodeConvNorm1(x)
        PorePressureControl = self.decoderConvTrans(PorePressureControl)
        PorePressureControl = self.decodeConvNorm1(PorePressureControl)
        StressControl = self.decoderConvTrans(StressControl)
        StressControl = self.decodeConvNorm1(StressControl)

        x = x + x2

        x = self.decoderConvTrans(x)
        x = self.decodeConvNorm1(x)
        PorePressureControl = self.decoderConvTrans(PorePressureControl)
        PorePressureControl = self.decodeConvNorm1(PorePressureControl)
        StressControl = self.decoderConvTrans(StressControl)
        StressControl = self.decodeConvNorm1(StressControl)

        feature = self.decoderConvTrans(feature)
        feature = self.decoderConvTrans(feature)

        temp = x * feature
        x = x + temp
        PorePressureControl = (PorePressureControl + PorePressureControl * feature + self.compute_Saturability_PorePressure(
                                      x)) / 2
        StressControl = StressControl + (StressControl * feature + x * PorePressureControl) / 2

        # 稳定性预测部分支路连接点
        Stabilization = (PorePressureControl + StressControl).detach()
        Stabilization = (Stabilization - Stabilization.mean()) / (Stabilization.std() + 1e-6)

        x = self.decoderConvTrans(x)
        x = self.decodeConvNorm1(x)
        PorePressureControl = self.decoderConvTrans(PorePressureControl)
        PorePressureControl = self.decodeConvNorm1(PorePressureControl)
        StressControl = self.decoderConvTrans(StressControl)
        StressControl = self.decodeConvNorm1(StressControl)

        x = self.final_conv(x)
        x = self.sigmoid(x)
        PorePressureControl = self.final_conv(PorePressureControl)
        PorePressureControl = self.sigmoid(PorePressureControl)
        StressControl = self.final_conv(StressControl)
        StressControl = self.sigmoid(StressControl)

        # 稳定性预测部分架构
        fs, fv = self.FeatureModuleStab(featureinit.float())

        Stabilization = self.stabilyBasicModule(Stabilization)

        Stabilization = Stabilization + fs

        Stabilization = self.flatten(Stabilization)

        Stabilization = F.relu(self.fc1(Stabilization))

        Stabilization = self.fc2(Stabilization) + self.flatten(fv)
        Stabilization = F.relu(Stabilization)
        Stabilization = self.fc3(Stabilization)

        return x, PorePressureControl, StressControl, Stabilization

# 定义设备
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
# 损失函数
criterion = nn.MSELoss()

# 自定义损失函数
def modelLoss(net, XTrain1, XTrain2, XTrain3, XTrain4, Y1, Y2, Y3, Y5, epoch, i):
    XTrain1 = XTrain1.to(device)
    XTrain2 = XTrain2.to(device)
    XTrain3 = XTrain3.to(device)
    XTrain4 = XTrain4.to(device)
    Y1 = Y1.to(device)
    Y2 = Y2.to(device)
    Y3 = Y3.to(device)
    Y5 = Y5.to(device)

    # 图像和特征的前向传播
    outputs1, outputs2, outputs3, outputs5 = net(XTrain1, XTrain2, XTrain3, XTrain4)
    # 在指定 epoch 保存预测图像
    if epoch is not None and epoch % 10 == 0 and i is not None:
        output_img = outputs1[0].detach().cpu().numpy().transpose(1, 2, 0)
        output_img = (output_img * 255).astype(np.uint8)
        save_dir = "MutilControl训练过程图像/饱和度StabilityUncertaintyWeighting"
        os.makedirs(save_dir, exist_ok=True)
        save_path = os.path.join(save_dir, f"Train_epoch_{epoch + 1}_{i + 1}.png")
        plt.imsave(save_path, output_img)

        output_img = outputs2[0].detach().cpu().numpy().transpose(1, 2, 0)
        output_img = (output_img * 255).astype(np.uint8)
        save_dir = "MutilControl训练过程图像/孔隙压力StabilityUncertaintyWeighting"
        os.makedirs(save_dir, exist_ok=True)
        save_path = os.path.join(save_dir, f"Train_epoch_{epoch + 1}_{i + 1}.png")
        plt.imsave(save_path, output_img)

        output_img = outputs3[0].detach().cpu().numpy().transpose(1, 2, 0)
        output_img = (output_img * 255).astype(np.uint8)
        save_dir = "MutilControl训练过程图像/应力StabilityUncertaintyWeighting"
        os.makedirs(save_dir, exist_ok=True)
        save_path = os.path.join(save_dir, f"Train_epoch_{epoch + 1}_{i + 1}.png")
        plt.imsave(save_path, output_img)

        with open('MutilControl训练过程图像/BaseUncertaintyWeightingtrain_results.txt','a') as file:
            file.write(f'Train_epoch_{epoch + 1}_{i + 1}_{outputs5}\n')


    # 数据驱动数值计算损失
    loss1 = criterion(outputs1, Y1)
    loss2 = criterion(outputs2, Y2)
    loss3 = criterion(outputs3, Y3)
    loss5 = criterion(outputs5.view(-1).float(), Y5.float())

    # 动态调整损失权重
    loss = torch.exp(-net.log_var1) * (loss1 + loss2 + loss3) + torch.exp(-net.log_var2) * loss5

    # 计算梯度
    gradients = torch.autograd.grad(loss, net.parameters(), retain_graph=True, create_graph=True,allow_unused=True)

    return loss, loss1, loss2, loss3, loss5, gradients, {}
def compute_gradient(field, dx, dy):
    grad_x = (field[:, :, :, 2:] - field[:, :, :, :-2]) / (2 * dx)
    grad_y = (field[:, :, 2:, :] - field[:, :, :-2, :]) / (2 * dy)
    grad_x = F.pad(grad_x, (1, 1, 0, 0), mode='replicate')
    grad_y = F.pad(grad_y, (0, 0, 1, 1), mode='replicate')
    return grad_x, grad_y

def train_model(datasetTrain, model_save_dir='models', batch_size=4, epochs=500, lr=0.01):

    dataloader = DataLoader(datasetTrain, batch_size=batch_size, shuffle=True)
    # 初始化模型
    model = SlopeNet()
    model.to(device)

    # 优化器
    optimizer = torch.optim.Adam(list(model.parameters()), lr=lr)
    # 训练
    num_epochs = epochs
    model.train()
    for epoch in range(num_epochs):
        # feature_extractor.train()
        running_loss = 0.0
        for i, (x1, x2, x3, x4, y1, y2, y3, y5) in enumerate(dataloader):
            x1, x2, x3, x4, y1, y2, y3, y5 = (x1.to(device), x2.unsqueeze(1).to(device),
                                                  x3.to(device), x4.to(device),
                                                  y1.to(device), y2.to(device), y3.to(device),
                                                  y5.to(device))

            # 计算损失
            loss, loss1, loss2, loss3, loss5, gradients, _ = modelLoss(model, x1, x2, x3, x4, y1, y2, y3, y5, epoch, i)

            # 反向传播
            optimizer.zero_grad()
            loss.backward()
            optimizer.step()

            running_loss += loss.item()
            print(f"Epoch [{epoch + 1}/{num_epochs}] Batch [{i + 1}],"
                  f" Loss: {loss.item():.4f}, Loss1: {loss1.item():.4f}, Loss2: {loss2.item():.4f},"
                  f" Loss3: {loss3.item():.4f}, Loss5: {loss5.item():.4f}")

        print(f"Epoch [{epoch+1}/{num_epochs}], Loss: {running_loss / len(dataloader):.4f}")
        if (epoch + 1) % 100 == 0:
            # 模型训练好之后的保存路径
            snapshot_path = model_save_dir
            os.makedirs(snapshot_path, exist_ok=True)
            save_mode_path = os.path.join(snapshot_path,
                                          'NonPLossBaselineMutilModelUnetStability_epoch_' +
                                          str(epoch + 1) + '.pth')
            torch.save(model.state_dict(), save_mode_path)

def value_to_Fr(outputs):
    """
    将 outputs 中的值四舍五入为整数索引，并从 Fr 数组中提取对应值。

    参数:
        outputs (np.ndarray 或 torch.Tensor): 输入的浮点数数组/张量

    返回:
        outputs1 (np.ndarray): 从 Fr 中提取的对应值
    """
    # 定义 Fr 数组（固定范围 0.98:0.00005:1.12）
    Fr = np.arange(0.98, 1.1200001, 0.00005)  # 2801 个点 1.1200001 确保包含 1.12

    # 处理输入类型（兼容 NumPy 和 PyTorch）
    if hasattr(outputs, 'numpy'):  # 如果是 PyTorch 张量
        outputs_np = outputs.detach().cpu().numpy()  # 转为 NumPy 数组
    else:
        outputs_np = np.array(outputs)  # 确保是 NumPy 数组

    # 1. 四舍五入取整，并限制索引范围
    indices = np.round(outputs_np).astype(int)
    indices = np.clip(indices, 0, len(Fr) - 1)  # 避免越界

    # 2. 从 Fr 中提取值
    outputs1 = Fr[indices]

    return outputs1

def rtest_model(datasetTest, InfiltrationStrength, model_path, output_dir='output', batch_size=1):

    dataloaderTest = DataLoader(datasetTest, batch_size=batch_size, shuffle=False)
    # 加载模型
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    model = SlopeNet().to(device)
    model.load_state_dict(torch.load(model_path, weights_only=True, map_location=device))
    model.eval()

    # 图像质量指标
    psnr_metric = PeakSignalNoiseRatio().to(device)
    ssim_metric = StructuralSimilarityIndexMeasure(data_range=1.0).to(device)

    # 回归任务指标
    mse_metric = MeanSquaredError().to(device)
    mae_metric = MeanAbsoluteError().to(device)
    r2_metric = R2Score().to(device)
    # 测试
    model.eval()
    test_loss = 0.0

    # 初始化指标存储
    metrics = {
        'task1': {'psnr': [], 'ssim': []},
        'task2': {'psnr': [], 'ssim': []},
        'task3': {'psnr': [], 'ssim': []},
        'task5': {'mse': [], 'mae': [], 'preds': [], 'trues': []}  # 改为收集所有预测和真实值
    }

    with ((torch.no_grad())):
        for i, (x1, x2, x3, x4, y1, y2, y3, y5 ) in enumerate(dataloaderTest):
            x1, x2, x3, x4, y1, y2, y3, y5 = (x1.to(device), x2.unsqueeze(1).to(device),
                                                  x3.to(device), x4.to(device),
                                                  y1.to(device), y2.to(device), y3.to(device),
                                                  y5.to(device))
            outputs1, outputs2, outputs3, outputs5 = model(x1, x2, x3, x4)

            loss = criterion(outputs1, y1) + criterion(outputs2, y2) + criterion(outputs3, y3) + criterion(outputs5.squeeze(0), y5)
            test_loss += loss.item()

            # 图像任务评估
            for out, gt, task in zip(
                    [outputs1, outputs2, outputs3],
                    [y1, y2, y3],
                    ['task1', 'task2', 'task3']
            ):
                # 计算PSNR和SSIM
                metrics[task]['psnr'].append(psnr_metric(out, gt))
                metrics[task]['ssim'].append(ssim_metric(out, gt))

            output_img = outputs1[0].detach().cpu().numpy().transpose(1, 2, 0)
            output_img = (output_img * 255).astype(np.uint8)
            save_dir = f"{output_dir}/NonPLossBaseline饱和度StabilityUncertaintyWeighting"
            os.makedirs(save_dir, exist_ok=True)
            save_path = os.path.join(save_dir, f"prediction_{InfiltrationStrength[i]}_{(Time[i] + 1)}.png")
            plt.imsave(save_path, output_img)

            output_img = outputs2[0].detach().cpu().numpy().transpose(1, 2, 0)
            output_img = (output_img * 255).astype(np.uint8)
            save_dir = f"{output_dir}/NonPLossBaseline孔隙压力StabilityUncertaintyWeighting"
            os.makedirs(save_dir, exist_ok=True)
            save_path = os.path.join(save_dir, f"prediction_{InfiltrationStrength[i]}_{(Time[i] + 1)}.png")
            plt.imsave(save_path, output_img)

            output_img = outputs3[0].detach().cpu().numpy().transpose(1, 2, 0)
            output_img = (output_img * 255).astype(np.uint8)
            save_dir = f"{output_dir}/NonPLossBaseline应力StabilityUncertaintyWeighting"
            os.makedirs(save_dir, exist_ok=True)
            save_path = os.path.join(save_dir, f"prediction_{InfiltrationStrength[i]}_{(Time[i] + 1)}.png")
            plt.imsave(save_path, output_img)

            # 处理数值预测任务
            outputs5 = torch.tensor(np.round(value_to_Fr(outputs5), 5))
            y5 = torch.tensor(np.round(value_to_Fr(y5), 5))

            y5_pred = outputs5.item()
            y5_true = y5.item()
            print(f"Predicted: {y5_pred}, True: {y5_true}")
            # 数值任务评估
            outputs5 = outputs5.squeeze(0)
            metrics['task5']['mse'].append(mse_metric(outputs5.float(), y5.float()))
            metrics['task5']['mae'].append(mae_metric(outputs5.float(), y5.float()))
            metrics['task5']['preds'].append(outputs5.float())
            metrics['task5']['trues'].append(y5.float())
            # 检查文件是否存在，不存在则新建，存在则加载
            output_file = f"{output_dir}/NonPLossBaselineUncertaintyWeighting_prediction_results.xlsx"
            if os.path.exists(output_file):
                wb = load_workbook(output_file)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                # 可选：添加表头
                ws.append(["InfiltrationStrength", "Time", "Prediction", "GroundTruth"])

            # 写入数据
            ws.append([InfiltrationStrength[i], Time[i] + 1, outputs5.item(), y5.item()])

            # 保存文件
            wb.save(output_file)

    # 计算平均指标
    # 计算最终R2分数（在所有样本上计算）
    if len(metrics['task5']['preds']) >= 2:
        all_preds = torch.stack(metrics['task5']['preds'])
        all_trues = torch.stack(metrics['task5']['trues'])

        # 确保计算时使用float类型
        mean_trues = torch.mean(all_trues.float())  # 显式转换为float
        ss_res = torch.sum((all_trues - all_preds) ** 2)
        ss_tot = torch.sum((all_trues - mean_trues) ** 2)
        r2_score = 1 - ss_res / (ss_tot + 1e-10)  # 添加小量防止除以0
        metrics['task5']['r2'] = r2_score.item()
    else:
        metrics['task5']['r2'] = float('nan')
        print("Warning: Not enough samples to compute R2 score (need at least 2)")

    # 1. 将 metrics 转换为 DataFrame
    def save_metrics_to_excel(metrics, output_path="metrics_results.xlsx"):
        # 深拷贝metrics并转换所有CUDA张量
        metrics_cpu = copy.deepcopy(metrics)

        for task in metrics_cpu.values():
            for key, values in task.items():
                if isinstance(values, list) and len(values) > 0 and hasattr(values[0], 'is_cuda'):
                    task[key] = [v.cpu().numpy() if hasattr(v, 'is_cuda') else v for v in values]

        # 创建一个Excel写入器
        writer = pd.ExcelWriter(output_path, engine='openpyxl')

        # 为每个task创建一个Sheet
        for task_name, task_metrics in metrics_cpu.items():
            # 将字典转换为DataFrame
            df = pd.DataFrame(task_metrics)

            # 添加统计信息（均值、标准差等）
            if task_name != 'task5':  # 图像任务（PSNR/SSIM）
                stats = df.agg(['mean', 'std', 'max', 'min']).T
            else:  # 数值任务（MSE/MAE）
                stats = df[['mse', 'mae']].agg(['mean', 'std', 'max', 'min']).T

            # 合并数据和统计信息
            combined_df = pd.concat([df, stats], axis=0)

            # 写入Excel的Sheet
            combined_df.to_excel(writer, sheet_name=task_name, index=True)

        # 保存Excel文件
        writer.close()
        print(f"Metrics saved to {output_path}")

    # 调用函数保存（替换为你的实际输出路径）
    save_metrics_to_excel(metrics, output_dir + "/NonPLossBaselinemetrics_results_summary.xlsx")

    # 计算平均指标
    def compute_avg_metrics(metrics):
        avg_metrics = {}
        for task in metrics:
            avg_metrics[task] = {}
            for metric in metrics[task]:
                if metric == 'r2':  # R2已经计算过
                    avg_metrics[task][metric] = metrics[task][metric]
                elif metrics[task][metric]:
                    if isinstance(metrics[task][metric][0], torch.Tensor):
                        avg_metrics[task][metric] = torch.stack(metrics[task][metric]).mean().item()
                    else:
                        avg_metrics[task][metric] = torch.tensor(metrics[task][metric]).mean().item()
        return avg_metrics

    avg_metrics = compute_avg_metrics(metrics)

    # 打印结果
    print("\n=== Final Metrics ===")
    for task in avg_metrics:
        print(f"\n{task.upper()}:")
        for metric, value in avg_metrics[task].items():
            print(f"{metric.upper()}: {value:.4f}")

    # 保存结果
    with open(f'{output_dir}/NonPLossBaselineResults.txt', 'w') as f:
        for task in avg_metrics:
            f.write(f"\n{task.upper()}:\n")
            for metric, value in avg_metrics[task].items():
                f.write(f"{metric}: {value:.4f}\n")

if __name__ == '__main__':

    # 训练数据路径
    file_path0025 = '../dataset/0025.xlsx'
    file_path0075 = '../dataset/0075.xlsx'
    file_path01 = '../dataset/0100.xlsx'
    file_path015 = '../dataset/0150.xlsx'
    file_path02 = '../dataset/0200.xlsx'
    data0025 = pd.read_excel(file_path0025, sheet_name='Sheet2')
    data0075 = pd.read_excel(file_path0075, sheet_name='Sheet2')
    data0100 = pd.read_excel(file_path01, sheet_name='Sheet2')
    data015 = pd.read_excel(file_path015, sheet_name='Sheet2')
    data02 = pd.read_excel(file_path02, sheet_name='Sheet2')
    Time = np.concatenate((data0100['TIME'].array, data015['TIME'].array, data02['TIME'].array,
                           data0025['TIME'].array, data0075['TIME'].array))
    InfiltrationStrength = np.concatenate((data0100['CLASS'].array, data015['CLASS'].array, data02['CLASS'].array,
                                           data0025['CLASS'].array, data0075['CLASS'].array))
    StabilizationFactor = np.concatenate((data0100['StabilizationFactor'].array, data015['StabilizationFactor'].array,
                                          data02['StabilizationFactor'].array, data0025['StabilizationFactor'].array,
                                          data0075['StabilizationFactor'].array))
    # 数据集路径
    x1_path = "../dataset/SaturabilityTraincopy"
    x21_values = InfiltrationStrength  # 降雨等级
    x22_values = Time  # 入渗时间或降雨时长
    x2_values = np.column_stack((x21_values, x22_values))
    x3_path = "../dataset/PorePressureControlTraincopy"
    x4_path = "../dataset/StressControlTraincopy"

    y1_path = "../dataset/Saturabilitycopy"
    y2_path = "../dataset/PorePressureControlcopy"
    y3_path = "../dataset/StressControlcopy"
    y5_values = StabilizationFactor  # 稳定性
    # 加载训练数据以及数据集划分和批次设置
    datasetTrain = SlopeDataset(x1_path, x2_values, x3_path, x4_path, y1_path, y2_path, y3_path, y5_values,
                           transform=transform)

    # 训练模型
    # train_history = train_model(
    #     datasetTrain=datasetTrain,
    #     model_save_dir="model_snapshot",
    #     batch_size=4,
    #     epochs=100,
    #     lr=0.001
    # )

    # 测试数据路径
    file_path005 = '../dataset/0050/0050.xlsx'
    # file_path0125 = '../dataset/0125/0125.xlsx'
    # file_path0175 = '../dataset/0175/0175.xlsx'
    data = pd.read_excel(file_path005, sheet_name='Sheet2')
    # data = pd.read_excel(file_path0125, sheet_name='Sheet2')
    # data = pd.read_excel(file_path0175, sheet_name='Sheet2')
    Time = data['TIME'].array
    InfiltrationStrength = data['CLASS'].array
    StabilizationFactor = data['StabilizationFactor'].array
    # 测试数据集路径
    name = '0050'
    x1_path = f"../dataset/{name}/SaturabilityTraincopy"
    x21_values = InfiltrationStrength  # 降雨等级
    x22_values = Time  # 入渗时间或降雨时长
    x2_values = np.column_stack((x21_values, x22_values))
    x3_path = f"../dataset/{name}/PorePressureControlTraincopy"
    x4_path = f"../dataset/{name}/StressControlTraincopy"

    y1_path = f"../dataset/{name}/Saturabilitycopy"
    y2_path = f"../dataset/{name}/PorePressureControlcopy"
    y3_path = f"../dataset/{name}/StressControlcopy"
    y5_values = StabilizationFactor  # 稳定性
    # 加载测试数据以及批次设置
    datasetTest = SlopeDataset(x1_path, x2_values, x3_path, x4_path, y1_path, y2_path, y3_path, y5_values,
                           transform=transform)

    # 测试模型
    rtest_model(
        datasetTest=datasetTest,
        InfiltrationStrength=InfiltrationStrength,
        model_path='model_snapshot/NonPLossBaselineMutilModelUnetStability_epoch_100.pth',
        output_dir='MutilControl结果图像0050'
    )
